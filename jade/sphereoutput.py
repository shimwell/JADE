# -*- coding: utf-8 -*-
"""
Created on Thu Jan  2 10:36:38 2020

@author: Davide Laghi

Copyright 2021, the JADE Development Team. All rights reserved.

This file is part of JADE.

JADE is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

JADE is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with JADE.  If not, see <http://www.gnu.org/licenses/>.
"""
from __future__ import annotations

import math
import os
import shutil
import sys

from typing import TYPE_CHECKING

import numpy as np
import pandas as pd
import openpyxl

from tqdm import tqdm
from xlsxwriter.utility import xl_rowcol_to_cell

import jade.atlas as at
import jade.excelsupport as exsupp
import jade.plotter as plotter
from jade.configuration import Configuration
from jade.output import BenchmarkOutput, MCNPoutput, OpenMCOutput

if TYPE_CHECKING:
    from jade.main import Session


class SphereOutput(BenchmarkOutput):
    def __init__(self, lib: str, code: str, testname: str, session: Session):
        """
        Initialises the SphereOutput class from the general BenchmarkOutput
        class, see output.py for details on how self variables are assigned

        Parameters
        ----------
        lib : str
            library to post-process
        code : str
            code being post processed
        testname : str
            name of the benchmark being postprocessed
        session : Session
            Jade Session
        exp : str
            the benchmark is an experimental one

        Returns
        -------
        None.

        """
        super().__init__(lib, code, testname, session)

        # Load the settings for zaids and materials
        mat_path = os.path.join(self.cnf_path, "MaterialsSettings.csv")
        self.mat_settings = pd.read_csv(mat_path, sep=",").set_index("Symbol")

        zaid_path = os.path.join(self.cnf_path, "ZaidSettings.csv")
        self.zaid_settings = pd.read_csv(zaid_path, sep=",").set_index("Z")

    def single_postprocess(self):
        """
        Execute the full post-processing of a single library (i.e. excel,
        raw data and atlas)

        Returns
        -------
        None.

        """
        print(" Generating Excel Recap...")
        self.pp_excel_single()
        print(" Dumping Raw Data...")
        self.print_raw()
        print(" Generating plots...")
        self._generate_single_plots()

    def _generate_single_plots(self):
        """
        Generate all the requested plots in a temporary folder

        Returns
        -------
        None.

        """

        for code, outputs in self.outputs.items():
            # edited by T. Wheeler. openmc requires separate tally numbers which is accounted for here
            outpath = os.path.join(self.atlas_path, "tmp")
            os.mkdir(outpath)
            if self.openmc:
                tally_info = [
                    (
                        4,
                        "Averaged Neutron Flux (175 groups)",
                        "Neutron Flux",
                        r"$\#/cm^2$",
                    ),
                    (14, "Averaged Gamma Flux (24 groups)", "Gamma Flux", r"$\#/cm^2$"),
                ]
            else:
                tally_info = [
                    (
                        2,
                        "Averaged Neutron Flux (175 groups)",
                        "Neutron Flux",
                        r"$\#/cm^2$",
                    ),
                    (32, "Averaged Gamma Flux (24 groups)", "Gamma Flux", r"$\#/cm^2$"),
                ]
            for tally, title, quantity, unit in tally_info:
                print(" Plotting tally n." + str(tally))
                for zaidnum, output in tqdm(outputs.items()):
                    title = title
                    tally_data = output.tallydata.set_index("Tally N.").loc[tally]
                    energy = tally_data["Energy"].values
                    values = tally_data["Value"].values
                    error = tally_data["Error"].values
                    lib_name = self.session.conf.get_lib_name(self.lib)
                    lib = {
                        "x": energy,
                        "y": values,
                        "err": error,
                        "ylabel": str(zaidnum) + " (" + lib_name + ")",
                    }
                    data = [lib]
                    outname = str(zaidnum) + "-" + self.lib + "-" + str(tally)
                    plot = plotter.Plotter(
                        data,
                        title,
                        outpath,
                        outname,
                        quantity,
                        unit,
                        "Energy [MeV]",
                        self.testname,
                    )
                    plot.plot("Binned graph")

            self._build_atlas(outpath)

    def _build_atlas(self, outpath):
        """
        Build the atlas using all plots contained in directory

        Parameters
        ----------
        outpath : str or path
            temporary folder containing all plots.

        Returns
        -------
        None.

        """
        atlas_path = os.path.join(outpath, "..")
        # Printing Atlas
        template = os.path.join(self.path_templates, "AtlasTemplate.docx")
        if self.single:
            name = self.lib
        else:
            name = self.name

        atlas = at.Atlas(template, "Sphere " + name)
        atlas.build(outpath, self.session.lib_manager, self.mat_settings)
        atlas.save(atlas_path)
        # Remove tmp images
        shutil.rmtree(outpath)

    def compare(self):
        """
        Execute the full post-processing of a comparison of libraries
        (i.e. excel, and atlas)

        Returns
        -------
        None.

        """
        print(" Generating Excel Recap...")
        self.pp_excel_comparison()
        print(" Creating Atlas...")
        # outpath = os.path.join(self.atlas_path, 'tmp')
        # os.mkdir(outpath)
        # Recover all libraries and zaids involved
        libraries, allzaids, outputs = self._get_organized_output()

        globalname = ""
        for lib in self.lib:
            globalname = globalname + lib + "_Vs_"

        globalname = globalname[:-4]

        # Plot everything
        print(" Generating Plots Atlas...")
        self._generate_plots(allzaids, globalname)

    def _generate_plots(self, allzaids, globalname):
        """
        Generate all the plots requested by the Sphere leakage benchmark

        Parameters
        ----------
        allzaids : list
            list of all zaids resulting from the union of the results from
            both libraries.
        outputs : dic
            dictionary containing the outputs for each library, for each code
            format: {
                    code1:{library1:[outputs], library2:[outputs], ...},
                    code2:{library1:[outputs], library2:[outputs], ...},
                    ...
                    }
        globalname : str
            name for the output.

        Returns
        -------
        None.

        """
        for code, code_outputs in self.outputs.items():
            outpath = os.path.join(self.atlas_path, "tmp")
            if not os.path.exists(outpath):
                os.mkdir(outpath)
            if code == "mcnp":
                tally_info = [
                    (
                        2,
                        "Averaged Neutron Flux (175 groups)",
                        "Neutron Flux",
                        r"$\#/cm^2$",
                    ),
                    (32, "Averaged Gamma Flux (24 groups)", "Gamma Flux", r"$\#/cm^2$"),
                ]
            if code == "openmc":
                tally_info = [
                    (
                        4,
                        "Averaged Neutron Flux (175 groups)",
                        "Neutron Flux",
                        r"$\#/cm^2$",
                    ),
                    (14, "Averaged Gamma Flux (24 groups)", "Gamma Flux", r"$\#/cm^2$"),
                ]
            for tally, title, quantity, unit in tally_info:
                print(" Plotting tally n." + str(tally))
                for zaidnum in tqdm(allzaids):
                    # title = title
                    data = []
                    for library, lib_outputs in code_outputs.items():
                        try:  # Zaid could not be common to the libraries
                            tally_data = (
                                lib_outputs[zaidnum]
                                .tallydata.set_index("Tally N.")
                                .loc[tally]
                            )
                            # print(lib_outputs[zaidnum])
                            energy = tally_data["Energy"].values
                            values = tally_data["Value"].values
                            error = tally_data["Error"].values
                            lib_name = self.session.conf.get_lib_name(library)
                            lib = {
                                "x": energy,
                                "y": values,
                                "err": error,
                                "ylabel": str(zaidnum) + " (" + str(lib_name) + ")",
                            }
                            data.append(lib)
                        except KeyError:
                            # It is ok, simply nothing to plot here
                            pass

                    outname = str(zaidnum) + "-" + globalname + "-" + str(tally)
                    plot = plotter.Plotter(
                        data,
                        title,
                        outpath,
                        outname,
                        quantity,
                        unit,
                        "Energy [MeV]",
                        self.testname,
                    )
                    try:
                        plot.plot("Binned graph")
                    except IndexError:
                        print(data)

            self._build_atlas(outpath)

    def _get_organized_output(self):
        """
        Organizes the outputs for each library in each code in the
        outputs object

        Returns:
        libraries: list
            list of all libraries to be post processed
        allzaids: list
            list of all zaids/materials that have been run
        outputs: list
            list of all output objects for all codes and all libraries

        """
        libraries = []
        outputs = []
        zaids = []
        for code, library_outputs in self.outputs.items():
            for libname, outputslib in library_outputs.items():
                libraries.append(libname)
                outputs.append(outputslib)
                zaids.append(list(outputslib.keys()))
        # Extend list to all zaids
        allzaids = zaids[0]
        for zaidlist in zaids[1:]:
            allzaids.extend(zaidlist)
        allzaids = set(allzaids)  # no duplicates

        return libraries, allzaids, outputs

    def _read_mcnp_output(self):
        """Reads all MCNP outputs from a library

        Returns
        -------
            outputs : dic
                Dictionary of MCNP sphere output objects used in plotting, keys are material name or ZAID number
            results : dic
                Dictionary of overview of Tally values for each material/ZAID, returns either all values > 0 for
                tallies with postiive values only, all Values = 0 for empty tallies, and returns the corresponding
                tally bin if it finds any negative values. Contents of the "Values" worksheet.
            errors : dic
                Dictionary of average errors for each tally for each material/Zaid. Contents of the "Errors" worksheet.
            stat_checks : dic
                Dictionary the MCNP statistical check results for each material/ZAID. Contents of the "Statistical
                Checks" Worksheet.
        """
        # Get results
        results = []
        errors = []
        stat_checks = []
        outputs = {}
        # test_path_mcnp = os.path.join(self.test_path, "mcnp")
        for folder in sorted(os.listdir(self.test_path)):
            results_path = os.path.join(self.test_path, folder, "mcnp")
            pieces = folder.split("_")
            # Get zaid
            zaidnum = pieces[-2]
            # Check for material exception
            if zaidnum == "Sphere":
                zaidnum = pieces[-1].upper()
                zaidname = self.mat_settings.loc[zaidnum, "Name"]
            else:
                zaidname = pieces[-1]
            # Get mfile
            for file in os.listdir(results_path):
                if file[-1] == "m":
                    mfile = file
                elif file[-1] == "o":
                    ofile = file
            # Parse output
            output = SphereMCNPoutput(
                os.path.join(results_path, mfile), os.path.join(results_path, ofile)
            )
            outputs[zaidnum] = output
            # Adjourn raw Data
            self.raw_data["mcnp"][zaidnum] = output.tallydata
            # Recover statistical checks
            st_ck = output.stat_checks
            # Recover results and precisions
            res, err = output.get_single_excel_data(
                ["2", "4", "6", "12", "14", "24", "34", "22", "32", "44", "46"]
            )
            for dic in [res, err, st_ck]:
                dic["Zaid"] = zaidnum
                dic["Zaid/Mat Name"] = zaidname
            results.append(res)
            errors.append(err)
            stat_checks.append(st_ck)
        return outputs, results, errors, stat_checks

    def _read_serpent_output(self):
        """Reads all Serpent outputs from a library

        NOT YET IMPLEMENTED

        Returns
        -------
            outputs : dic
                Dictionary of Serpent sphere output objects used in plotting, keys are material name or ZAID number
            results : dic
                Dictionary of overview of Tally values for each material/ZAID, returns either all values > 0 for
                tallies with postiive values only, all Values = 0 for empty tallies, and returns the corresponding
                tally bin if it finds any negative values. Contents of the "Values" worksheet.
            errors : dic
                Dictionary of average errors for each tally for each material/Zaid. Contents of the "Errors" worksheet.
        """
        # Get results
        results = []
        errors = []
        stat_checks = []
        outputs = {}
        test_path_serpent = os.path.join(self.test_path, "serpent")
        for folder in os.listdir(test_path_serpent):
            # Call parser here
            continue
        return outputs, results, errors, stat_checks

    def _read_openmc_output(self):
        """Reads all OpenMC outputs from a library

        Returns
        -------
            outputs : dic
                Dictionary of OpenMC sphere output objects used for plotting, keys are material name or ZAID number
            results : dic
                Dictionary of overview of Tally values for each material/ZAID, returns either all values > 0 for
                tallies with postiive values only, all Values = 0 for empty tallies, and returns the corresponding
                tally bin if it finds any negative values. Contents of the "Values" worksheet.
            errors : dic
                Dictionary of average errors for each tally for each material/Zaid. Contents of the "Errors" worksheet.
        """
        # Get results
        results = []
        errors = []
        # stat_checks = []
        outputs = {}
        # test_path_openmc = os.path.join(self.test_path, "openmc")
        for folder in sorted(os.listdir(self.test_path)):
            results_path = os.path.join(self.test_path, folder, "openmc")
            pieces = folder.split("_")
            # Get zaid
            zaidnum = pieces[-2]
            # Check for material exception
            if zaidnum == "Sphere":
                zaidnum = pieces[-1].upper()
                zaidname = self.mat_settings.loc[zaidnum, "Name"]
            else:
                zaidname = pieces[-1]
            # Parse output
            output = SphereOpenMCoutput(os.path.join(results_path, "tallies.out"))
            outputs[zaidnum] = output
            # Adjourn raw Data
            self.raw_data["openmc"][zaidnum] = output.tallydata
            # Recover statistical checks
            # st_ck = output.stat_checks
            # Recover results and precisions
            res, err = output.get_single_excel_data(["4", "14"])
            for dic in [res, err]:
                dic["Zaid"] = zaidnum
                dic["Zaid/Mat Name"] = zaidname
            results.append(res)
            errors.append(err)
            # stat_checks.append(st_ck)
        return (
            outputs,
            results,
            errors,
        )  # stat_checks

    def _generate_dataframe(self, results, errors, stat_checks=None):
        """Function to turn the output of the read_{code}_output functions into DataFrames
           for use with xlsxwriter

        Arguments
        ------
            results (dic): dictionary of tally summaries for each material/ZAID.
            errors (dic): dictionaty of average tally errors across all energy bins.
            stat_checks (dic, optional): dictionary containing results of MCNP statistical checks
            (MCNP only). Defaults to None.

        Returns
        -------
            results (DataFrame): previous dictionary but in DataFrame form
            errors (DataFrame): previous dictionary but in DataFrame form
            stat_checks (DataFrame): previous dictionary but in DataFrame form
        """
        # Generate DataFrames
        results = pd.DataFrame(results)
        errors = pd.DataFrame(errors)

        # Swap Columns and correct zaid sorting
        # results
        for df in [results, errors]:
            df["index"] = pd.to_numeric(df["Zaid"].values, errors="coerce")
            df.sort_values("index", inplace=True)
            del df["index"]

            df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
            df.reset_index(inplace=True)

        if stat_checks is not None:
            stat_checks = pd.DataFrame(stat_checks)
            stat_checks["index"] = pd.to_numeric(
                stat_checks["Zaid"].values, errors="coerce"
            )
            stat_checks.sort_values("index", inplace=True)
            del stat_checks["index"]

            stat_checks.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
            stat_checks.reset_index(inplace=True)
        return results, errors, stat_checks

    def pp_excel_single(self):
        """
        Generate the single library results excel

        Returns
        -------
        None.

        """
        self.outputs = {}
        self.results = {}
        self.errors = {}
        self.stat_checks = {}

        if self.mcnp:
            outfolder_path = self.excel_path
            # os.makedirs(outfolder_path, exist_ok=True)
            # outpath = os.path.join(self.excel_path_mcnp,'Sphere_single_' + 'MCNP_' + self.lib+'.xlsx')
            outpath = os.path.join(
                outfolder_path, "Sphere_single_" + "MCNP_" + self.lib + ".xlsx"
            )
            outputs, results, errors, stat_checks = self._read_mcnp_output()
            results, errors, stat_checks = self._generate_dataframe(
                results, errors, stat_checks
            )
            self.outputs["mcnp"] = outputs
            self.results["mcnp"] = results
            self.errors["mcnp"] = errors
            self.stat_checks["mcnp"] = stat_checks
            lib_name = self.session.conf.get_lib_name(self.lib)
            # Generate DataFrames
            # results = pd.DataFrame(results)
            # errors = pd.DataFrame(errors)
            # stat_checks = pd.DataFrame(stat_checks)

            # Swap Columns and correct zaid sorting
            # results
            # for df in [results, errors, stat_checks]:
            #    df['index'] = pd.to_numeric(df['Zaid'].values, errors='coerce')
            #    df.sort_values('index', inplace=True)
            #    del df['index']

            #    df.set_index(['Zaid', 'Zaid Name'], inplace=True)
            #    df.reset_index(inplace=True)
            exsupp.sphere_single_excel_writer(
                self, outpath, lib_name, results, errors, stat_checks
            )

        if self.serpent:
            pass

        if self.openmc:
            outfolder_path = self.excel_path
            # os.mkdir(outfolder_path)
            # outpath = os.path.join(self.excel_path_openmc,'Sphere_single_' + 'OpenMC_' + self.lib+'.xlsx')
            outpath = os.path.join(
                outfolder_path, "Sphere_single_" + "OpenMC_" + self.lib + ".xlsx"
            )
            outputs, results, errors = self._read_openmc_output()
            results, errors, stat_checks = self._generate_dataframe(results, errors)
            self.outputs["openmc"] = outputs
            self.results["openmc"] = results
            self.errors["openmc"] = errors
            self.stat_checks["openmc"] = stat_checks

            exsupp.sphere_single_excel_writer(self, outpath, self.lib, results, errors)

        if self.d1s:
            pass

    def pp_excel_comparison(self):
        """
        Compute the data and create the excel for all libraries comparisons.
        In the meantime, additional data is stored for future plots.


        Returns
        -------
        None.

        """

        code_outputs = {}

        if self.mcnp:
            iteration = 0
            outputs = {}
            for reflib, tarlib, name in self.couples:
                outfolder_path = self.excel_path
                # os.mkdir(outfolder_path)
                outpath = os.path.join(
                    outfolder_path, "Sphere_comparison_" + name + "_mcnp.xlsx"
                )
                # outpath = os.path.join(self.excel_path_mcnp, 'Sphere_comparison_' +
                #                       name+'.xlsx')
                # Get results
                comp_dfs = []
                error_dfs = []

                for test_path in [
                    self.test_path[reflib],
                    self.test_path[tarlib],
                ]:
                    results = []
                    errors = []
                    iteration = iteration + 1
                    outputs_lib = {}
                    for folder in os.listdir(test_path):
                        results_path = os.path.join(test_path, folder, "mcnp")
                        pieces = folder.split("_")
                        # Get zaid
                        zaidnum = pieces[-2]
                        # Check for material exception
                        if zaidnum == "Sphere":
                            zaidnum = pieces[-1].upper()
                            zaidname = self.mat_settings.loc[zaidnum, "Name"]
                        else:
                            zaidname = pieces[-1]

                        # Get mfile
                        for file in os.listdir(results_path):
                            if file[-1] == "m":
                                mfile = file
                            elif file[-1] == "o":
                                outfile = file

                        # Parse output
                        mfile = os.path.join(results_path, mfile)
                        outfile = os.path.join(results_path, outfile)
                        output = SphereMCNPoutput(mfile, outfile)
                        outputs_lib[zaidnum] = output
                        res, err, columns = output.get_comparison_data(
                            ["12", "22", "24", "14", "34", "6", "46"], "mcnp"
                        )
                        try:
                            zn = int(zaidnum)
                        except ValueError:  # Happens for typical materials
                            zn = zaidnum

                        res.append(zn)
                        err.append(zn)
                        res.append(zaidname)
                        err.append(zaidname)

                        results.append(res)
                        errors.append(err)

                    # Add reference library outputs
                    if iteration == 1:
                        outputs[reflib] = outputs_lib

                    if iteration == 2:
                        outputs[tarlib] = outputs_lib

                    # Generate DataFrames
                    columns.extend(["Zaid", "Zaid/Mat Name"])
                    comp_df = pd.DataFrame(results, columns=columns)
                    error_df = pd.DataFrame(errors, columns=columns)
                    comp_df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
                    error_df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
                    comp_dfs.append(comp_df)
                    error_dfs.append(error_df)

                code_outputs["mcnp"] = outputs
                self.outputs = code_outputs
                # self.results["mcnp"] = results
                # self.errors["mcnp"] = errors

                # Consider only common zaids
                idx1 = comp_dfs[0].index
                idx2 = comp_dfs[1].index
                newidx = idx1.intersection(idx2)

                # Build the final excel data
                final = (comp_dfs[0].loc[newidx] - comp_dfs[1].loc[newidx]) / comp_dfs[
                    0
                ].loc[newidx]
                absdiff = comp_dfs[0].loc[newidx] - comp_dfs[1].loc[newidx]

                # self.diff_data["mcnp"] = final
                # self.absdiff["mcnp"] = absdiff

                # Standard deviation
                idx1 = absdiff.index
                idx2 = error_dfs[0].index
                newidx = idx1.intersection(idx2)

                std_dev = absdiff.loc[newidx] / error_dfs[0].loc[newidx]

                # self.std_dev["mcnp"] = std_dev
                # Correct sorting
                for df in [final, absdiff, std_dev]:
                    df.reset_index(inplace=True)
                    df["index"] = pd.to_numeric(df["Zaid"].values, errors="coerce")
                    df.sort_values("index", inplace=True)
                    del df["index"]
                    df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)

                # Create and concat the summary
                old_l = 0
                old_lim = 0
                rows = []
                limits = [0, 0.05, 0.1, 0.2, 0.2]
                for i, sup_lim in enumerate(limits[1:]):
                    if i == len(limits) - 2:
                        row = {"Range": "% of cells > " + str(sup_lim * 100)}
                        for column in final.columns:
                            cleaned = final[column].replace("", np.nan).dropna()
                            l_range = len(cleaned[abs(cleaned) > sup_lim])
                            try:
                                row[column] = l_range / len(cleaned)
                            except ZeroDivisionError:
                                row[column] = np.nan
                    else:
                        row = {
                            "Range": str(old_lim * 100)
                            + " < "
                            + "% of cells"
                            + " < "
                            + str(sup_lim * 100)
                        }
                        for column in final.columns:
                            cleaned = final[column].replace("", np.nan).dropna()
                            lenght = len(cleaned[abs(cleaned) < sup_lim])
                            old_l = len(cleaned[abs(cleaned) < limits[i]])
                            l_range = lenght - old_l
                            try:
                                row[column] = l_range / len(cleaned)
                            except ZeroDivisionError:
                                row[column] = np.nan

                    old_lim = sup_lim
                    rows.append(row)

                summary = pd.DataFrame(rows)
                summary.set_index("Range", inplace=True)
                # If it is zero the CS are equal! (NaN if both zeros)
                for df in [final, absdiff, std_dev]:
                    # df[df == np.nan] = 'Not Available'
                    df.astype({col: float for col in df.columns[1:]})
                    df.replace(np.nan, "Not Available", inplace=True)
                    df.replace(float(0), "Identical", inplace=True)
                    df.replace(-np.inf, "Reference = 0", inplace=True)
                    df.replace(1, "Target = 0", inplace=True)

                # retrieve single pp files to add as extra tabs to comparison workbook
                single_pp_files = []
                # Add single pp sheets
                for lib in [reflib, tarlib]:
                    pp_dir = self.session.state.get_path(
                        "single", [lib, "Sphere", "mcnp", "Excel"]
                    )
                    pp_file = os.listdir(pp_dir)[0]
                    single_pp_path = os.path.join(pp_dir, pp_file)
                    single_pp_files.append(single_pp_path)

                # --- Write excel ---
                # Generate the excel
                exsupp.sphere_comp_excel_writer(
                    self,
                    outpath,
                    name,
                    final,
                    absdiff,
                    std_dev,
                    summary,
                    single_pp_files,
                )

                # # Add single pp sheets
                # current_wb = openpyxl.load_workbook(outpath)
                # for lib in [reflib, tarlib]:
                #     cp = self.session.state.get_path(
                #         "single", [lib, "Sphere", "mcnp", "Excel"]
                #     )
                #     file = os.listdir(cp)[0]
                #     cp = os.path.join(cp, file)
                #     # open file
                #     single_wb = openpyxl.load_workbook(cp)
                #     for ws in single_wb.worksheets:
                #         destination = current_wb.create_sheet(ws.title + " " + lib)
                #         exsupp.copy_sheet(ws, destination)
                #     single_wb.close()

                # current_wb.save(outpath)
                # current_wb.close()

                # ex.save()
                # """
        if self.openmc:
            iteration = 0
            outputs = {}
            for reflib, tarlib, name in self.couples:
                outfolder_path = self.excel_path
                # os.mkdir(outfolder_path)
                outpath = os.path.join(
                    outfolder_path, "Sphere_comparison_" + name + "_openmc.xlsx"
                )
                # outpath = os.path.join(self.excel_path_openmc, 'Sphere_comparison_' +
                #                       name+'openmc.xlsx')
                # Get results
                comp_dfs = []
                error_dfs = []

                for test_path in [self.test_path[reflib], self.test_path[tarlib]]:
                    results = []
                    errors = []
                    iteration = iteration + 1
                    outputs_lib = {}
                    for folder in os.listdir(test_path):
                        results_path = os.path.join(test_path, folder, "openmc")
                        pieces = folder.split("_")
                        # Get zaid
                        zaidnum = pieces[-2]
                        # Check for material exception
                        if zaidnum == "Sphere":
                            zaidnum = pieces[-1].upper()
                            zaidname = self.mat_settings.loc[zaidnum, "Name"]
                        else:
                            zaidname = pieces[-1]

                        # Get mfile
                        for file in os.listdir(results_path):
                            if "tallies.out" in file:
                                outfile = file

                        # Parse output
                        outfile = os.path.join(results_path, outfile)
                        output = SphereOpenMCoutput(outfile)
                        outputs_lib[zaidnum] = output
                        res, err, columns = output.get_comparison_data(
                            ["4", "14"], "openmc"
                        )
                        try:
                            zn = int(zaidnum)
                        except ValueError:  # Happens for typical materials
                            zn = zaidnum

                        res.append(zn)
                        err.append(zn)
                        res.append(zaidname)
                        err.append(zaidname)

                        results.append(res)
                        errors.append(err)
                    # Add reference library outputs
                    if iteration == 1:
                        outputs[reflib] = outputs_lib

                    if test_path == os.path.join(self.test_path[tarlib], "openmc"):
                        outputs[tarlib] = outputs_lib

                    # Generate DataFrames
                    columns.extend(["Zaid", "Zaid/Mat Name"])
                    comp_df = pd.DataFrame(results, columns=columns)
                    error_df = pd.DataFrame(errors, columns=columns)
                    comp_df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
                    error_df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
                    comp_dfs.append(comp_df)
                    error_dfs.append(error_df)

                    # outputs_couple = outputs
                    # self.results = results
                code_outputs["openmc"] = outputs
                self.outputs = code_outputs
                # self.results["openmc"] = results
                # self.errors["openmc"] = errors
                # Consider only common zaids
                idx1 = comp_dfs[0].index
                idx2 = comp_dfs[1].index
                newidx = idx1.intersection(idx2)

                # Build the final excel data
                final = (comp_dfs[0].loc[newidx] - comp_dfs[1].loc[newidx]) / comp_dfs[
                    0
                ].loc[newidx]
                absdiff = comp_dfs[0].loc[newidx] - comp_dfs[1].loc[newidx]

                # self.diff_data["openmc"] = final
                # self.absdiff["openmc"] = absdiff

                # Standard deviation
                idx1 = absdiff.index
                idx2 = error_dfs[0].index
                newidx = idx1.intersection(idx2)

                std_dev = absdiff.loc[newidx] / error_dfs[0].loc[newidx]

                # self.std_dev["openmc"] = std_dev

                # Correct sorting
                for df in [final, absdiff, std_dev]:
                    df.reset_index(inplace=True)
                    df["index"] = pd.to_numeric(df["Zaid"].values, errors="coerce")
                    df.sort_values("index", inplace=True)
                    del df["index"]
                    df.set_index(["Zaid", "Zaid/Mat Name"], inplace=True)
                # Create and concat the summary
                old_l = 0
                old_lim = 0
                rows = []
                limits = [0, 0.05, 0.1, 0.2, 0.2]
                for i, sup_lim in enumerate(limits[1:]):
                    if i == len(limits) - 2:
                        row = {"Range": "% of cells > " + str(sup_lim * 100)}
                        for column in final.columns:
                            cleaned = final[column].replace("", np.nan).dropna()
                            l_range = len(cleaned[abs(cleaned) > sup_lim])
                            try:
                                row[column] = l_range / len(cleaned)
                            except ZeroDivisionError:
                                row[column] = np.nan
                    else:
                        row = {
                            "Range": str(old_lim * 100)
                            + " < "
                            + "% of cells"
                            + " < "
                            + str(sup_lim * 100)
                        }
                        for column in final.columns:
                            cleaned = final[column].replace("", np.nan).dropna()
                            lenght = len(cleaned[abs(cleaned) < sup_lim])
                            old_l = len(cleaned[abs(cleaned) < limits[i]])
                            l_range = lenght - old_l
                            try:
                                row[column] = l_range / len(cleaned)
                            except ZeroDivisionError:
                                row[column] = np.nan

                    old_lim = sup_lim
                    rows.append(row)

                summary = pd.DataFrame(rows)
                summary.set_index("Range", inplace=True)
                # If it is zero the CS are equal! (NaN if both zeros)
                for df in [final, absdiff, std_dev]:
                    # df[df == np.nan] = 'Not Available'
                    df.astype({col: float for col in df.columns[1:]})
                    df.replace(np.nan, "Not Available", inplace=True)
                    df.replace(float(0), "Identical", inplace=True)
                    df.replace(-np.inf, "Reference = 0", inplace=True)
                    df.replace(1, "Target = 0", inplace=True)

                # retrieve single pp files to add as extra tabs to comparison workbook
                single_pp_files = []
                # Add single pp sheets
                for lib in [reflib, tarlib]:
                    pp_dir = self.session.state.get_path(
                        "single", [lib, "Sphere", "openmc", "Excel"]
                    )
                    pp_file = os.listdir(pp_dir)[0]
                    single_pp_path = os.path.join(pp_dir, pp_file)
                    single_pp_files.append(single_pp_path)

                # --- Write excel ---
                # Generate the excel
                exsupp.sphere_comp_excel_writer(
                    self, outpath, name, final, absdiff, std_dev, summary, single_pp_files
                )
        if self.serpent:
            pass

    def print_raw(self):
        """
        Assigns a path and prints the post processing data as a .csv

        """
        if self.mcnp:
            for key, data in self.raw_data["mcnp"].items():
                file = os.path.join(self.raw_path, "mcnp" + key + ".csv")
                data.to_csv(file, header=True, index=False)
        if self.serpent:
            for key, data in self.raw_data["serpent"].items():
                file = os.path.join(self.raw_path, "serpent" + key + ".csv")
                data.to_csv(file, header=True, index=False)
        if self.openmc:
            for key, data in self.raw_data["openmc"].items():
                file = os.path.join(self.raw_path, "openmc" + key + ".csv")
                data.to_csv(file, header=True, index=False)
        if self.d1s:
            for key, data in self.raw_data["d1s"].items():
                file = os.path.join(self.raw_path, "d1s" + key + ".csv")
                data.to_csv(file, header=True, index=False)


class SphereTallyOutput:
    def get_single_excel_data(self, tallies2pp):
        """
        Get the excel data of a single MCNP output

        Returns
        -------
        results : dic
            Excel result for different tallies
        errors : dic
            Error average in all tallies

        """

        data = self.tallydata.set_index(["Energy"])
        results = {}  # Store excel results of different tallies
        errors = {}  # Store average error in different tallies
        heating_res = {}  # Mid-process heating results
        notes = "Negative Bins:"  # Record negative bins here
        initial_notes_length = len(notes)  # To check if notes are registered
        tally_list = [d for _, d in data.groupby(["Tally N."])]
        heating_tallies = ["4", "6", "44", "46"]
        for tally in tally_list:
            tally_num = str(tally["Tally N."].iloc[0])
            tally_description = tally["Tally Description"].iloc[0]
            mean_error = tally["Error"].mean()
            if tally_num in heating_tallies:
                heating_res[tally_num] = tally["Value"].values[0]
            if tally_num in tallies2pp:
                tally_zero = tally[tally["Value"] == 0]
                original_length = len(tally)
                tally = tally[tally["Value"] < 0]
                if len(tally) > 0:
                    res = "Value < 0 in " + str(len(tally)) + " bin(s)"
                    # Get energy bins
                    bins = list(tally.reset_index()["Energy"].values)
                    notes = notes + "\n(" + str(tally_num) + "): "
                    for ebin in bins:
                        notes = notes + str(ebin) + ", "
                    notes = notes[:-2]  # Clear string from excess commas
                elif len(tally_zero) == original_length:
                    res = "Value = 0 for all bins"
                else:
                    res = "Value > 0 for all bins"
                results[tally_description] = res
                errors[tally_description] = mean_error

        if len(heating_res) == 4:
            comp = "Heating comparison [F4 vs F6]"
            try:
                results["Neutron " + comp] = (
                    heating_res["6"] - heating_res["4"]
                ) / heating_res["6"]
            except ZeroDivisionError:
                results["Neutron " + comp] = 0

            try:
                results["Gamma " + comp] = (
                    heating_res["46"] - heating_res["44"]
                ) / heating_res["46"]
            except ZeroDivisionError:
                results["Gamma " + comp] = 0

        # Notes adding
        if len(notes) > initial_notes_length:
            results["Notes"] = notes
        else:
            results["Notes"] = ""

        return results, errors

    def get_comparison_data(self, tallies2pp, code):
        """
        Get Data for single zaid to be used in comparison.

        Returns
        -------
        results : list
            All results per tally to compare
        columns : list
            Tally names

        """
        # Tallies to post process
        if code == "mcnp":
            binned_tallies = ["12", "22"]
        if code == "openmc":
            binned_tallies = ["4", "14"]
        integral_tallies = []

        # Acquire data
        data = self.tallydata.set_index(["Energy"])
        totalbins = self.totalbin.set_index("Tally Description")
        results = []  # Store data to compare for different tallies
        errors = []
        columns = []  # Tally names and numbers

        # Reorder tallies
        tallies = {}
        tally_list = [d for _, d in data.groupby(["Tally N."])]
        for tally in tally_list:
            tally_str = str(tally["Tally N."].iloc[0])
            if tally_str in tallies2pp:
                tallies[tally_str] = tally
                if tally_str not in binned_tallies:
                    integral_tallies.append(tally_str)

        # Process binned tallies
        for tally_num in binned_tallies:
            tally = tallies[tally_num]
            tally_description = tally["Tally Description"].iloc[0]
            # Get energy bins
            bins = tally.index.tolist()
            for ebin in bins:
                colname = str(ebin) + " [MeV]" + " [t" + tally_num + "]"
                columns.append(colname)
                results.append(tally["Value"].loc[ebin])
                errors.append(tally["Error"].loc[ebin])
            # Add the total bin
            colname = "Total" + " [t" + tally_num + "]"
            columns.append(colname)
            results.append(totalbins["Value"].loc[tally_description])
            errors.append(totalbins["Error"].loc[tally_description])

        # Proccess integral tallies
        for tally_num in integral_tallies:
            tally = tallies[tally_num]
            tally_description = tally["Tally Description"].iloc[0]
            columns.append(tally_description)
            results.append(float(tally["Value"]))
            errors.append(float(tally["Error"]))
        return results, errors, columns


class SphereMCNPoutput(MCNPoutput, SphereTallyOutput):
    def organize_mctal(self):
        """
        Retrieve and organize mctal data. Simplified for sphere leakage case

        Returns: DataFrame containing the organized data
        """
        # Extract data
        rows = []
        rowstotal = []
        for t in self.mctal.tallies:
            num = t.tallyNumber
            des = t.tallyComment[0]
            nCells = t.getNbins("f", False)
            nCora = t.getNbins("i", False)
            nCorb = t.getNbins("j", False)
            nCorc = t.getNbins("k", False)
            nDir = t.getNbins("d", False)
            # usrAxis = t.getAxis("u")
            nUsr = t.getNbins("u", False)
            # segAxis = t.getAxis("s")
            nSeg = t.getNbins("s", False)
            nMul = t.getNbins("m", False)
            # cosAxis = t.getAxis("c")
            nCos = t.getNbins("c", False)
            # ergAxis = t.getAxis("e")
            nErg = t.getNbins("e", False)
            # timAxis = t.getAxis("t")
            nTim = t.getNbins("t", False)

            for f in range(nCells):
                for d in range(nDir):
                    for u in range(nUsr):
                        for s in range(nSeg):
                            for m in range(nMul):
                                for c in range(nCos):
                                    for e in range(nErg):
                                        try:
                                            erg = t.erg[e]
                                        except IndexError:
                                            erg = None

                                        for nt in range(nTim):
                                            for k in range(nCorc):
                                                for j in range(nCorb):
                                                    for i in range(nCora):
                                                        val = t.getValue(
                                                            f,
                                                            d,
                                                            u,
                                                            s,
                                                            m,
                                                            c,
                                                            e,
                                                            nt,
                                                            i,
                                                            j,
                                                            k,
                                                            0,
                                                        )
                                                        err = t.getValue(
                                                            f,
                                                            d,
                                                            u,
                                                            s,
                                                            m,
                                                            c,
                                                            e,
                                                            nt,
                                                            i,
                                                            j,
                                                            k,
                                                            1,
                                                        )
                                                        if val <= 0:
                                                            err = np.nan

                                                        row = [num, des, erg, val, err]
                                                        rows.append(row)

            # If Energy binning is involved
            if t.ergTC == "t":
                # 7 steps to get to energy, + 4 for time and mesh directions
                totalbin = t.valsErrors[-1][-1][-1][-1][-1][-1][-1][-1][-1][-1][-1]
                totalvalue = totalbin[0]
                if totalvalue > 0:
                    totalerror = totalbin[-1]
                else:
                    totalerror = np.nan
                row = [num, des, totalvalue, totalerror]
                rowstotal.append(row)

        df = pd.DataFrame(
            rows, columns=["Tally N.", "Tally Description", "Energy", "Value", "Error"]
        )
        dftotal = pd.DataFrame(
            rowstotal, columns=["Tally N.", "Tally Description", "Value", "Error"]
        )
        return df, dftotal


class SphereOpenMCoutput(OpenMCOutput, SphereTallyOutput):
    def _create_dataframe(self, rows):
        """Creates dataframe from the data in each output passed through as
        a list of lists from the process_tally function

        Args:
        rows: list
        list of list containing the rows of information from an output file

        Returns:
        df: DataFrame
        dataframe containing the information from each output

        dftotal: DataFrame
        dataframe containing the sum of all values and errors for each output
        """
        df = pd.DataFrame(
            rows, columns=["Tally N.", "Tally Description", "Energy", "Value", "Error"]
        )
        # rowstotal = [[rows[-1][0], rows[-1][1], 0.0, 0.0]]
        # for row in rows:
        #    rowstotal[0][2] += row[3]
        #    rowstotal[0][3] += row[4] ** 2
        # rowstotal[0][3] = np.sqrt(rowstotal[0][3])
        tallies = df["Tally N."].unique().tolist()
        descriptions = df["Tally Description"].unique().tolist()
        rowstotal = []
        for tally, description in zip(tallies, descriptions):
            value = df.loc[df["Tally N."] == tally, "Value"].sum()
            error = np.sqrt((df.loc[df["Tally N."] == tally, "Error"] ** 2).sum())
            rowstotal.append([tally, description, value, error])
        dftotal = pd.DataFrame(
            rowstotal, columns=["Tally N.", "Tally Description", "Value", "Error"]
        )
        return df, dftotal

    def process_tally(self):
        """
        Reads data from output file and stores it as a list of lists
        to be turned into a dataframe

        Returns:
            tallydata: Dataframe
            see df in _create_dataframe()

            totalbin: Dataframe
            see dftotal in _create_dataframe()
        """
        rows = []
        for line in self.output_file_data:
            if "tally" in line.lower():
                parts = line.split()
                tally_n = int(parts[2].replace(":", ""))
                tally_description = " ".join([parts[3].title(), parts[4].title()])
            if "incoming energy" in line.lower():
                parts = line.split()
                energy = 1e-6 * float(parts[3].replace(")", ""))
            if "flux" in line.lower():
                if ":" in line:
                    continue
                else:
                    parts = line.split()
                    value, error = float(parts[1]), float(parts[3])
                    rows.append([tally_n, tally_description, energy, value, error])
        tallydata, totalbin = self._create_dataframe(rows)
        return tallydata, totalbin


class SphereSDDRoutput(SphereOutput):
    times = ["0s", "2.7h", "24h", "11.6d", "30d", "10y"]
    timecols = {
        "0s": "1.0",
        "2.7h": "2.0",
        "24h": "3.0",
        "11.6d": "4.0",
        "30d": "5.0",
        "10y": "6.0",
    }

    def pp_excel_single(self):
        """
        Generate the single library results excel

        Returns
        -------
        None.

        """
        self.outputs = {}
        self.results = {}
        self.errors = {}
        self.stat_checks = {}
        if self.d1s:
            # template = os.path.join(os.getcwd(), "templates", "SphereSDDR_single.xlsx")
            outpath = os.path.join(
                self.excel_path, "SphereSDDR_single_" + self.lib + ".xlsx"
            )
            # compute the results
            outputs, results, errors, stat_checks = self._compute_single_results()
            self.outputs["d1s"] = outputs
            self.results["d1s"] = results
            self.errors["d1s"] = errors
            self.stat_checks["d1s"] = stat_checks
            lib_name = self.session.conf.get_lib_name(self.lib)
            # Write excel
            # ex = SphereExcelOutputSheet(template, outpath)
            # Results
            # ex.insert_df(11, 2, results, 0, header=False)
            # ex.insert_df(11, 2, errors, 1, header=False)
            # ex.insert_df(9, 2, stat_checks, 2, header=True)
            # lib_name = self.session.conf.get_lib_name(self.lib)
            # ex.wb.sheets[0].range("E1").value = lib_name
            # ex.save()
            exsupp.sphere_sddr_single_excel_writer(
                outpath, lib_name, results, errors, stat_checks
            )

    def pp_excel_comparison(self):
        """
        Generate the excel comparison output

        Returns
        -------
        None.

        """
        # template = os.path.join(os.getcwd(), "templates", "SphereSDDR_comparison.xlsx")
        if self.d1s:
            for reflib, tarlib, name in self.couples:
                outpath = os.path.join(
                    self.excel_path, "Sphere_SDDR_comparison_" + name + ".xlsx"
                )
                final, absdiff, std_dev = self._compute_compare_result(reflib, tarlib)

                # --- Write excel ---
                # Generate the excel
                # ex = SphereExcelOutputSheet(template, outpath)
                # Prepare the copy of the comparison sheet
                # ws_comp = ex.wb.sheets["Comparison"]
                # ws_diff = ex.wb.sheets["Comparison (Abs diff)"]

                # WRITE RESULTS
                # Percentage comparison
                # rangeex = ws_comp.range("B11")
                # rangeex.options(index=True, header=False).value = final
                # ws_comp.range("E1").value = name

                # Absolute difference comparison
                # rangeex = ws_diff.range("B11")
                # rangeex.options(index=True, header=False).value = absdiff

                single_pp_files = []
                # Add single pp sheets
                for lib in [reflib, tarlib]:
                    pp_dir = self.session.state.get_path(
                        "single", [lib, "SphereSDDR", "d1s", "Excel"]
                    )
                    pp_file = os.listdir(pp_dir)[0]
                    single_pp_path = os.path.join(pp_dir, pp_file)
                    single_pp_files.append(single_pp_path)

                exsupp.sphere_sddr_comp_excel_writer(
                    outpath, name, final, absdiff, std_dev, single_pp_files
                )

    def _get_organized_output(self):
        """
        Simply recover a list of the zaids and libraries involved
        """
        zaids = []
        for code, library_outputs in self.outputs.items():
            for (zaidnum, mt, lib), outputslib in library_outputs.items():
                zaids.append((zaidnum, mt))

        zaids = list(set(zaids))
        libs = []  # Not used
        outputs = []  # Not used

        return libs, zaids, outputs

    def _generate_single_plots(self):
        libs, allzaids, outputs = self._get_organized_output()
        globalname = self.lib
        self._generate_plots(allzaids, globalname)

    def _generate_plots(self, allzaids, globalname):
        """
        Generate all the plots requested by the Sphere SDDR benchmark

        Parameters
        ----------
        allzaids : list
            list of all zaids resulting from the union of the results from
            both libraries.
        globalname : str
            name for the output.

        Returns
        -------
        None.

        """
        # Check if self libraries is already a list
        if type(self.lib) != list:
            libraries = [self.lib]
        else:
            libraries = self.lib

        # Initialize atlas
        outpath = os.path.join(self.atlas_path, "tmp")
        if not os.path.exists(outpath):
            os.mkdir(outpath)
        template = os.path.join(self.path_templates, "AtlasTemplate.docx")
        atlas = at.Atlas(template, "Sphere SDDR " + globalname)
        libmanager = self.session.lib_manager

        # ------------- Binned plots of gamma flux ------------
        atlas.doc.add_heading("Photon Flux (32)", level=1)
        fluxquantity = "Photon Flux"
        fluxunit = r"$p/(cm^2\cdot\#_S)$"
        allzaids.sort()
        # --- Binned plots of the gamma flux ---
        for zaidnum, mt in tqdm(allzaids, desc=" Binned flux plots"):
            # Get everything for the title of the zaid
            try:
                name, formula = libmanager.get_zaidname(zaidnum)
                args = [zaidnum, name, formula, mt]
                title = "Zaid: {} ({} {}), MT={}".format(*args)
                # For zaids cooldown time does not change anything
                # Keep the multiple times only for materials
                times = [self.times[0]]
            except ValueError:  # A material is passed instead of zaid
                matname = self.mat_settings.loc[zaidnum, "Name"]
                title = zaidnum + " (" + matname + ")"
                times = self.times
            atlas.doc.add_heading(title, level=2)

            for time in times:
                atlas.doc.add_heading("Cooldown time = {}".format(time), level=3)
                title = "Gamma Leakage flux after a {} cooldown".format(time)
                data = []
                for lib in libraries:

                    try:  # Zaid could not be common to the libraries
                        outp = self.outputs["d1s"][zaidnum, mt, lib]
                    except KeyError:
                        # It is ok, simply nothing to plot here since zaid was
                        # not in library
                        continue
                    # Get the zaid flux
                    tally_data = outp.tallydata[32].set_index("Time")

                    # Select the correct time
                    t = "F" + self.timecols[time]
                    tally_data = tally_data.loc[t]
                    # If for some reason a total survived just kill him
                    tally_data = tally_data[tally_data.Energy != "total"]

                    energy = tally_data["Energy"].values
                    values = tally_data["Value"].values
                    error = tally_data["Error"].values
                    lib_name = self.session.conf.get_lib_name(lib)
                    ylabel = "{}_{} ({})".format(formula, mt, lib_name)
                    libdata = {"x": energy, "y": values, "err": error, "ylabel": ylabel}
                    data.append(libdata)

                outname = "{}-{}-{}-{}-{}".format(zaidnum, mt, globalname, 32, t)
                plot = plotter.Plotter(
                    data,
                    title,
                    outpath,
                    outname,
                    fluxquantity,
                    fluxunit,
                    "Energy [MeV]",
                    self.testname,
                )
                outfile = plot.plot("Binned graph")
                atlas.insert_img(outfile)

        # --- Wave plots flux ---
        # Do this block only if libs are more than one
        lim = 35  # limit of zaids to be in a single plot
        # Plot parameters which are not going to change
        quantity = ["Neutron Flux", "Photon Flux", "SDDR"]
        unit = [r"$n/(cm^2\cdot n_S)$", r"$p/(cm^2\cdot n_S)$", "Sv/h"]
        xlabel = "Zaid/Material and MT value"
        if len(libraries) > 1:
            atlas.doc.add_heading("Flux and SDDR ratio plots", level=1)
            # 1) collect zaid-mt couples in libraries and keep only the ones
            #    that appears on the reference + at least one lib
            # Build a df will all possible zaid, mt, lib combination
            if self.d1s:
                allkeys = list(self.outputs["d1s"].keys())
            df = pd.DataFrame(allkeys)
            df.columns = ["zaid", "mt", "lib"]
            df["zaid-mt"] = df["zaid"].astype(str) + "-" + df["mt"].astype(str)
            df.set_index("lib", inplace=True)
            # get the reference zaids
            refzaids = set(df.loc[self.lib[0]]["zaid-mt"].values)
            otherzaids = set(df.drop(self.lib[0])["zaid-mt"].values)
            # Get the final zaid-mt couples to consider
            zaid_couples = []
            mat_couples = []
            for zaidmt in refzaids:
                if zaidmt in otherzaids:
                    zaid, mt = zaidmt.split("-")
                    if zaid[0] in "mM":
                        mat_couples.append((zaid, mt))
                    else:
                        zaid_couples.append((zaid, mt))
            # sort it
            zaid_couples.sort(key=self._sortfunc_zaidMTcouples)
            mat_couples.sort(key=self._sortfunc_zaidMTcouples)

            # # There is going to be a plot for each cooldown time
            # Only one plot necessary for zaids at cd=0

            # 2) Recover/compute the data that needs to be plot for each lib
            data = []
            time = self.times[0]
            for lib in self.lib:
                nfluxs = []
                pfluxs = []
                sddrs = []
                xlabels = []
                ylabel = self.session.conf.get_lib_name(lib)
                for zaid, mt in zaid_couples:
                    # Extract values
                    nflux, pflux, sddr = self._extract_data4plots(zaid, mt, lib, time)
                    # Memorize values
                    nfluxs.append(nflux)
                    pfluxs.append(pflux)
                    sddrs.append(sddr)
                    name, formula = libmanager.get_zaidname(zaid)
                    xlabels.append(formula + " " + mt)

                # Split the data if its length is more then the limit
                datalenght = len(xlabels)
                sets = math.ceil(datalenght / lim)
                last_idx = 0
                idxs = []
                step = int(datalenght / sets)
                for _ in range(sets):
                    newidx = last_idx + step
                    idxs.append((last_idx, newidx))
                    last_idx = newidx

                for j, (start, end) in enumerate(idxs):
                    # build the dic
                    ydata = [nfluxs[start:end], pfluxs[start:end], sddrs[start:end]]
                    xlab = xlabels[start:end]
                    libdata = {"x": xlab, "y": ydata, "err": [], "ylabel": ylabel}
                    # try to append it to the data in the correct index
                    # if the index is not found, then the list still needs
                    # to be initialized
                    try:
                        data[j].append(libdata)
                    except IndexError:
                        data.append([libdata])

            # 3) Compute parameters for the plotter init
            refname = self.session.conf.get_lib_name(self.lib[0])
            for datapiece in data:
                title = "Ratio Vs {} (T0 + {})".format(refname, time)
                outname = "dummy"  # Does not matter if plot is added imm.
                testname = self.testname
                plot = plotter.Plotter(
                    datapiece, title, outpath, outname, quantity, unit, xlabel, testname
                )
                outfile = plot.plot("Waves")
                atlas.insert_img(outfile)

            # --- Single wave plot for each material ---
            atlas.doc.add_heading("Materials ratio plot", level=1)
            xlab = self.times
            quantity = ["Neutron Flux", "Photon Flux", "SDDR"]
            unit = ["", "", ""]
            xlabel = "Cooldown time"
            for material, _ in tqdm(mat_couples, desc=" Materials: "):
                atlas.doc.add_heading(material, level=2)
                data = []
                for lib in self.lib:
                    ylabel = self.session.conf.get_lib_name(lib)
                    nfluxs = []
                    pfluxs = []
                    sddrs = []
                    for time in self.times:
                        nflux, pflux, sddr = self._extract_data4plots(
                            material, "All", lib, time
                        )
                        # Memorize
                        nfluxs.append(nflux)
                        pfluxs.append(pflux)
                        sddrs.append(sddr)

                    # Build the lib data
                    ydata = [nfluxs, pfluxs, sddrs]

                    libdata = {"x": xlab, "y": ydata, "err": [], "ylabel": ylabel}
                    data.append(libdata)

                # Plot
                refname = self.session.conf.get_lib_name(self.lib[0])
                matname = self.mat_settings.loc[material, "Name"]
                title = "Ratio Vs {} ({})".format(refname, matname)
                outname = "dummy"  # Does not matter if plot is added imm.
                testname = self.testname
                plot = plotter.Plotter(
                    data, title, outpath, outname, quantity, unit, xlabel, testname
                )
                outfile = plot.plot("Waves")
                atlas.insert_img(outfile)

        ########
        print(" Building...")
        if self.d1s:
            atlas.save(self.atlas_path)
        # Remove tmp images
        shutil.rmtree(outpath)

    def _extract_data4plots(self, zaid, mt, lib, time):
        """_summary_

        Args:
            zaid (str): zaid of output
            mt (str): mt
            lib (str): library being postprocessed
            time (float): timestep

        Returns:
            nflux (float): neutron flux
            pflux (float): proton flux
            sddr (float): shut down dose rate
        """
        if self.d1s:
            tallies = self.outputs["d1s"][zaid, mt, lib].tallydata
        # Extract values
        nflux = tallies[12].set_index("Energy").drop("total")
        nflux = nflux.sum().loc["Value"]
        pflux = tallies[22].groupby("Time").sum().loc[1, "Value"]
        sddr = tallies[104].set_index("Time")
        sddr = sddr.loc["D" + self.timecols[time], "Value"]
        # Memorize values
        return nflux, pflux, sddr

    def _compute_single_results(
        self,
    ) -> tuple[dict[str, dict], pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Compute the excel single post processing results and memorize them

        Parameters
        ----------

        Returns
        -------
        outputs: dict[str, dict]
            dictionary of the outputs. the first level is the code level
        results : pd.DataFrame
            global excel datataframe of all values.
        errors : pd.DataFrame
            global excel dataframe of all errors.
        stat_checks : pd.DataFrame
            global excel dataframe of all statistical checks.

        """
        # Get results
        # results = []
        # errors = []
        # stat_checks = []
        desc = " Parsing Outputs: "
        if self.d1s:
            # for folder in tqdm(os.listdir(self.test_path), desc=desc):
            outputs, results, errors, stat_checks = self._parserunmcnp(
                self.test_path, self.lib
            )

            self.outputs["d1s"] = outputs

        # Generate DataFrames
        results = pd.concat(results, axis=1).T
        errors = pd.concat(errors, axis=1).T
        stat_checks = pd.DataFrame(stat_checks)

        # Swap Columns and correct zaid sorting
        # results
        for df in [results, errors, stat_checks]:
            self._sort_df(df)  # it is sorted in place
            df.set_index("Parent")

        # self.outputs = outputs

        return outputs, results, errors, stat_checks

    def _compute_compare_result(self, reflib, tarlib):
        """
        Given a reference lib and a target lib, both absolute and relative
        comparison are computed

        Parameters
        ----------
        reflib : str
            reference library suffix.
        tarlib : str
            target library suffix.

        Returns
        -------
        final : pd.DataFrame
            relative comparison table.
        absdiff : pd.DataFrame
            absolute comparison table.
        std_dev:
            comparison in std. dev. from mean table

        """
        # Get results both of the reflib and tarlib
        comp_dfs = []
        error_dfs = []
        lib_dics = []
        code_outputs = {}
        test_paths = [self.test_path[reflib], self.test_path[tarlib]]
        libs = [reflib, tarlib]
        for test_path, lib in zip(test_paths, libs):
            # Extract all the series from the different reactions
            # Collect the data
            outputs, results, errors, _ = self._parserunmcnp(test_path, lib)
            # Build the df and sort
            comp_df = pd.concat(results, axis=1).T
            error_df = pd.concat(errors, axis=1).T
            for df in [comp_df, error_df]:
                self._sort_df(df)
                # They need to be indexed
                df.set_index(["Parent", "Parent Name", "MT"], inplace=True)
                # Add the df to the list
            comp_dfs.append(comp_df)
            error_dfs.append(error_df)
            lib_dics.append(outputs)
        for dic in lib_dics:
            code_outputs.update(dic)
        self.outputs["d1s"] = code_outputs
        # Consider only common zaids
        idx1 = comp_dfs[0].index
        idx2 = comp_dfs[1].index
        newidx = idx1.intersection(idx2)
        # For some reason they arrive here as objects triggering
        # a ZeroDivisionError
        ref = comp_dfs[0].loc[newidx].astype(float)
        tar = comp_dfs[1].loc[newidx].astype(float)
        ref_err = error_dfs[1].loc[newidx].astype(float)

        # Build the final excel data
        absdiff = ref - tar
        final = absdiff / ref
        std_dev = absdiff / ref_err

        # If it is zero the CS are equal! (NaN if both zeros)
        for df in [final, absdiff, std_dev]:
            df.replace(np.nan, "Not Available", inplace=True)
            df.replace(float(0), "Identical", inplace=True)
            df.replace(-np.inf, "Reference = 0", inplace=True)
            df.replace(1, "Target = 0", inplace=True)

        return final, absdiff, std_dev

    @staticmethod
    def _sort_df(df):
        """
        Sorts the values in the passed dataframe by the Parent column,
        then sets 3 index columns

        Args:
            df (DataFrame): Dataframe containing output data
        """
        df["index"] = pd.to_numeric(df["Parent"].values, errors="coerce")
        df.sort_values("index", inplace=True)
        del df["index"]

        df.set_index(["Parent", "Parent Name", "MT"], inplace=True)
        df.reset_index(inplace=True)

    @staticmethod
    def _sortfunc_zaidMTcouples(item):
        try:
            zaid = int(item[0])
        except ValueError:
            zaid = item[0]
        try:
            mt = int(item[1])
        except ValueError:
            mt = item[1]

        if isinstance(zaid, str):
            flag = True
        else:
            flag = False

        return (flag, zaid, mt)

    def _parserunmcnp(self, test_path, lib):
        """
        given a MCNP run folder the parsing of the different outputs is
        performed

        Parameters
        ----------
        test_path : path or str
            path to the test.
        folder : str
            name of the folder to parse inside test_path.
        lib : str
            library.

        Returns
        -------
        outputs : Dictionary
            MCNP output object
        results : List
            List of results dataframes
        errors : List
            List of errors dataframes
        stat_checks : List
            List of stat checks dataframes


        """
        results = []
        errors = []
        stat_checks = []
        outputs = {}
        for folder in os.listdir(test_path):
            results_path = os.path.join(test_path, folder, "d1s")
            pieces = folder.split("_")
            # Get zaid
            zaidnum = pieces[1]
            # Check for material exception
            try:
                zaidname = self.mat_settings.loc[zaidnum, "Name"]
                mt = "All"
            except KeyError:
                # it is a simple zaid
                zaidname = pieces[2]
                mt = pieces[3]
            # Get mfile
            for file in os.listdir(results_path):
                if file[-1] == "m":
                    mfile = file
                elif file[-1] == "o":
                    ofile = file
                # Parse output
            output = SphereSDDRMCNPoutput(
                os.path.join(results_path, mfile), os.path.join(results_path, ofile)
            )

            outputs[zaidnum, mt, lib] = output
            # Adjourn raw Data
            self.raw_data["d1s"][zaidnum, mt, lib] = output.tallydata
            # Recover statistical checks
            st_ck = output.stat_checks
            # Recover results and precisions
            res, err = output.get_single_excel_data()
            for series in [res, err, st_ck]:
                series["Parent"] = zaidnum
                series["Parent Name"] = zaidname
                series["MT"] = mt
            results.append(res)
            errors.append(err)
            stat_checks.append(st_ck)

        return outputs, results, errors, stat_checks

    def print_raw(self):
        """
        Assigns a path and prints the post processing data as a .csv

        """
        if self.d1s:
            for key, data in self.raw_data["d1s"].items():
                foldername = "{}_{}".format(key[0], key[1])
                folder = os.path.join(self.raw_path, foldername)
                os.mkdir(folder)
                # Dump all tallies
                for tallynum, df in data.items():
                    filename = "{}_{}_{}.csv".format(key[0], key[1], tallynum)
                    file = os.path.join(self.raw_path, folder, filename)
                    df.to_csv(file, header=True, index=False)


class SphereSDDRMCNPoutput(SphereMCNPoutput):
    def organize_mctal(self):
        """
        Reorganize the MCTAL data in dataframes

        Returns
        -------
        tallydata : dic of DataFrame
            contains the tally data in a df format.
        totalbin : dic of DataFrame
            contain the total bin data.
        """
        # This should use the original MCNPotput organization of
        # MCTAL
        tallydata, totalbin = super(SphereMCNPoutput, self).organize_mctal()

        return tallydata, totalbin

    def get_single_excel_data(self):
        """
        Return the data that will be used in the single
        post-processing excel output for a single reaction

        Returns
        -------
        vals : pd.Series
            series reporting the result of a single reaction.
        errors : pd.Series
            series containing the errors associated with the
            reactions.

        """
        # 32 -> fine gamma flux
        # 104 -> Dose rate
        nflux = self.tallydata[12]
        pflux = self.tallydata[32]
        sddr = self.tallydata[104]
        heat = self.tallydata[46]

        # Differentiate time labels
        pflux["Time"] = "F" + pflux["Time"].astype(str)
        sddr["Time"] = "D" + sddr["Time"].astype(str)
        heat["Time"] = "H" + heat["Time"].astype(str)

        # Get the total values of the flux at different cooling times
        pfluxvals = pflux.groupby("Time").sum()["Value"]
        # Get the mean error of the flux at different cooling times
        pfluxerrors = pflux.groupby("Time").mean()["Error"]

        # Get the total values of the SDDR at different cooling times
        sddrvals = sddr.groupby("Time").sum()["Value"]
        # Get the mean error of the SDDR at different cooling times
        sddrerrors = sddr.groupby("Time").mean()["Error"]

        # Get the total Heating at different cooling times
        heatvals = heat.set_index("Time")["Value"]
        # Get the Heating mean error at different cooling times
        heaterrors = heat.set_index("Time")["Error"]

        # Neutron flux binned in energy
        nfluxvals = nflux.set_index("Energy")["Value"]
        # Errors of the neutron flux
        nfluxerrors = nflux.set_index("Energy")["Error"]

        # Delete the total row in case it is there
        for df, tag in zip(
            [pfluxvals, pfluxerrors, sddrvals, sddrerrors, heatvals, heaterrors],
            ["F", "F", "D", "D", "H", "H"],
        ):
            try:
                del df[tag + "total"]
            except KeyError:
                # If total value is not there it is ok
                pass

        # Do the same for the flux
        for df in [nfluxvals, nfluxerrors]:
            try:
                del df["total"]
            except KeyError:
                # If total value is not there it is ok
                pass

        # 2 series need to be built here, one for values and one for errors
        vals = pd.concat([pfluxvals, sddrvals, heatvals, nfluxvals], axis=0)
        errors = pd.concat([pfluxerrors, sddrerrors, heaterrors, nfluxerrors], axis=0)

        return vals, errors
